from fastapi import FastAPI, UploadFile, File, HTTPException, Body, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from parser_utils import parse_p4_structure
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tempfile
from typing import Dict, Any
from pathlib import Path

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="P4Lens API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


# Cleanup function for background tasks
def cleanup_file(path: str):
    """Delete file in background after response is sent."""
    try:
        if os.path.exists(path):
            os.remove(path)
            logger.info(f"Cleaned up file: {path}")
    except Exception as e:
        logger.error(f"Failed to cleanup {path}: {e}")


@app.get("/")
async def root():
    return {"message": "P4Lens API is running", "version": "1.0.0"}


@app.get("/health")
async def health():
    return {"status": "healthy"}


@app.post("/upload")
async def upload_p4(
    file: UploadFile = File(...), background_tasks: BackgroundTasks = None
):
    """Upload and parse P4 file."""
    # Validate file extension
    if not file.filename.endswith(".p4"):
        raise HTTPException(
            status_code=400, detail="Invalid file type. Please upload a .p4 file."
        )

    # Generate unique filename to avoid conflicts
    file_path = os.path.join(UPLOAD_DIR, f"temp_{os.getpid()}_{file.filename}")

    try:
        # Read and validate file content
        content = await file.read()

        if not content:
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")

        # Save file temporarily
        with open(file_path, "wb") as f:
            f.write(content)

        logger.info(f"Processing P4 file: {file.filename}")

        # Parse structure
        structure = parse_p4_structure(file_path)

        if (
            not structure
            or len([k for k in structure.keys() if not k.startswith("_")]) == 0
        ):
            raise HTTPException(
                status_code=400,
                detail="Could not parse P4 structure. File may be invalid or empty.",
            )

        logger.info(f"Successfully parsed {file.filename}")

        return {"filename": file.filename, "structure": structure}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing {file.filename}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error parsing P4 file: {str(e)}")
    finally:
        # Always cleanup the uploaded file
        if background_tasks:
            background_tasks.add_task(cleanup_file, file_path)
        else:
            cleanup_file(file_path)


def create_excel_export(structure: dict, filename: str) -> str:
    """Create Excel file with P4 rules and installation tables."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Define styles
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Sheet 1: Tables Overview
    ws_tables = wb.create_sheet("Tables & Rules")
    ws_tables.append(["P4 Tables and Match-Action Rules"])
    ws_tables.merge_cells("A1:F1")
    ws_tables["A1"].font = title_font
    ws_tables["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_tables.append([])
    headers = [
        "Table Name",
        "Match Keys",
        "Match Type",
        "Actions",
        "Size",
        "Default Action",
    ]
    ws_tables.append(headers)

    # Style header row
    for col in range(1, len(headers) + 1):
        cell = ws_tables.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    tables = structure.get("_tables", {})
    for table_name, table_info in tables.items():
        keys = table_info.get("keys", [])
        actions = table_info.get("actions", [])
        size = table_info.get("size", "N/A")
        default_action = table_info.get("default_action", "N/A")

        # Parse match types from keys
        match_types = []
        key_names = []
        for key in keys:
            if ":" in key:
                parts = key.split(":")
                key_names.append(parts[0].strip())
                match_types.append(parts[1].strip() if len(parts) > 1 else "exact")
            else:
                key_names.append(key.strip())
                match_types.append("exact")

        row = [
            table_name,
            "\n".join(key_names) if key_names else "None",
            "\n".join(match_types) if match_types else "N/A",
            "\n".join(actions) if actions else "None",
            size,
            default_action,
        ]
        ws_tables.append(row)

        # Add borders to data rows
        for col in range(1, len(row) + 1):
            cell = ws_tables.cell(row=ws_tables.max_row, column=col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Adjust column widths
    ws_tables.column_dimensions["A"].width = 20
    ws_tables.column_dimensions["B"].width = 30
    ws_tables.column_dimensions["C"].width = 15
    ws_tables.column_dimensions["D"].width = 25
    ws_tables.column_dimensions["E"].width = 10
    ws_tables.column_dimensions["F"].width = 20

    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()

    return temp_file.name


@app.post("/export-excel")
async def export_excel(
    structure: Dict[str, Any] = Body(...), background_tasks: BackgroundTasks = None
):
    """Export P4 structure to Excel format."""
    try:
        filename = structure.get("_filename", "p4_export")
        excel_path = create_excel_export(structure, filename)

        # Schedule cleanup after response is sent
        if background_tasks:
            background_tasks.add_task(cleanup_file, excel_path)

        return FileResponse(
            excel_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{filename.replace('.p4', '')}_rules.xlsx",
            background=background_tasks,
        )
    except Exception as e:
        logger.error(f"Error creating Excel export: {str(e)}")
        raise HTTPException(
            status_code=500, detail=f"Error creating Excel export: {str(e)}"
        )
